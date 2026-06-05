from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..database import Dividend, Trade
from .quotes import market_of

TRADE_HEADERS = ["type", "ticker", "shares", "price", "date", "fee", "notes", "market"]
DIVIDEND_HEADERS = ["ticker", "amount", "date", "notes", "market"]
TRADES_SHEET = "Trades"
DIVIDENDS_SHEET = "Dividends"


def portfolio_to_xlsx(
    trades: Iterable[Trade], dividends: Iterable[Dividend]
) -> bytes:
    """Serialize the portfolio to a two-sheet .xlsx workbook.

    Sheet ``Trades`` and sheet ``Dividends`` together hold everything the
    app needs to fully reconstruct its display — all derived numbers
    (P/L, holdings, charts) are recomputed from these on import.
    """
    wb = Workbook()

    ws_t = wb.active
    ws_t.title = TRADES_SHEET
    ws_t.append(TRADE_HEADERS)
    for t in trades:
        ws_t.append(
            [
                t.type,
                t.ticker,
                t.shares,
                t.price,
                t.trade_date,
                t.fee,
                t.notes or "",
                t.market or market_of(t.ticker),
            ]
        )

    ws_d = wb.create_sheet(DIVIDENDS_SHEET)
    ws_d.append(DIVIDEND_HEADERS)
    for d in dividends:
        ws_d.append(
            [
                d.ticker,
                d.amount,
                d.pay_date,
                d.notes or "",
                d.market or market_of(d.ticker),
            ]
        )

    for ws, headers, date_col in (
        (ws_t, TRADE_HEADERS, "E"),
        (ws_d, DIVIDEND_HEADERS, "C"),
    ):
        ws.freeze_panes = "A2"
        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 16
        for cell in ws[date_col][1:]:
            cell.number_format = "yyyy-mm-dd"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_date(value: object, where: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{where}: unrecognized date {s!r}")


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _cell_float(value: object, where: str, field: str) -> float:
    s = _cell_str(value)
    if s == "":
        raise ValueError(f"{where}: missing required value '{field}'")
    try:
        return float(s)
    except ValueError:
        raise ValueError(f"{where}: invalid number for '{field}': {s!r}")


def _header_map(row: tuple, sheet: str) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, cell in enumerate(row):
        name = _cell_str(cell).lower()
        if name:
            out[name] = idx
    if not out:
        raise ValueError(f"Sheet '{sheet}' has no header row")
    return out


def _row_is_blank(row: tuple) -> bool:
    return not any(_cell_str(c) for c in row)


def parse_portfolio_xlsx(data: bytes) -> tuple[list[Trade], list[Dividend]]:
    """Parse a two-sheet portfolio workbook into Trade / Dividend records.

    Trades come back sorted by (date asc, buy-before-sell) — the canonical
    order avg-cost / FIFO calculations expect, so a re-imported workbook
    reproduces the same realized P/L as before.
    """
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"Not a readable .xlsx file: {e}")

    sheets = {name.lower(): name for name in wb.sheetnames}
    if TRADES_SHEET.lower() not in sheets and DIVIDENDS_SHEET.lower() not in sheets:
        raise ValueError(
            "Workbook must contain a 'Trades' and/or 'Dividends' sheet"
        )

    trades: list[Trade] = []
    dividends: list[Dividend] = []

    if TRADES_SHEET.lower() in sheets:
        ws = wb[sheets[TRADES_SHEET.lower()]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = _header_map(next(rows), TRADES_SHEET)
        except StopIteration:
            header = {}
        required = {"type", "ticker", "shares", "price", "date"}
        missing = required - set(header)
        if header and missing:
            raise ValueError(
                f"Sheet 'Trades' is missing columns: {sorted(missing)}"
            )
        for i, row in enumerate(rows, start=2):
            if _row_is_blank(row):
                continue
            where = f"Trades row {i}"

            def get(col: str) -> object:
                idx = header.get(col)
                return row[idx] if idx is not None and idx < len(row) else None

            type_ = _cell_str(get("type")).lower()
            if type_ not in ("buy", "sell"):
                raise ValueError(
                    f"{where}: type must be 'buy' or 'sell', got {type_!r}"
                )
            ticker = _cell_str(get("ticker")).upper()
            if not ticker:
                raise ValueError(f"{where}: missing ticker")
            shares = _cell_float(get("shares"), where, "shares")
            price = _cell_float(get("price"), where, "price")
            if shares <= 0 or price <= 0:
                raise ValueError(f"{where}: shares and price must be > 0")
            fee_raw = _cell_str(get("fee"))
            fee = float(fee_raw) if fee_raw else 0.0
            if fee < 0:
                raise ValueError(f"{where}: fee must be >= 0")
            notes = _cell_str(get("notes")) or None
            # Optional column — old workbooks (pre-US support) won't have it, so
            # fall back to inferring the market from the ticker.
            market = _cell_str(get("market")).upper() or market_of(ticker)
            trades.append(
                Trade(
                    type=type_,
                    ticker=ticker,
                    shares=shares,
                    price=price,
                    trade_date=_parse_date(get("date"), where),
                    fee=fee,
                    notes=notes,
                    market=market,
                )
            )

    if DIVIDENDS_SHEET.lower() in sheets:
        ws = wb[sheets[DIVIDENDS_SHEET.lower()]]
        rows = ws.iter_rows(values_only=True)
        try:
            header = _header_map(next(rows), DIVIDENDS_SHEET)
        except StopIteration:
            header = {}
        required = {"ticker", "amount", "date"}
        missing = required - set(header)
        if header and missing:
            raise ValueError(
                f"Sheet 'Dividends' is missing columns: {sorted(missing)}"
            )
        for i, row in enumerate(rows, start=2):
            if _row_is_blank(row):
                continue
            where = f"Dividends row {i}"

            def get(col: str) -> object:
                idx = header.get(col)
                return row[idx] if idx is not None and idx < len(row) else None

            ticker = _cell_str(get("ticker")).upper()
            if not ticker:
                raise ValueError(f"{where}: missing ticker")
            amount = _cell_float(get("amount"), where, "amount")
            if amount <= 0:
                raise ValueError(f"{where}: amount must be > 0")
            notes = _cell_str(get("notes")) or None
            market = _cell_str(get("market")).upper() or market_of(ticker)
            dividends.append(
                Dividend(
                    ticker=ticker,
                    amount=amount,
                    pay_date=_parse_date(get("date"), where),
                    notes=notes,
                    market=market,
                )
            )

    wb.close()

    trades.sort(key=lambda t: (t.trade_date, 0 if t.type == "buy" else 1))
    dividends.sort(key=lambda d: d.pay_date)
    return trades, dividends


def insert_trades(db: Session, trades: list[Trade]) -> int:
    for t in trades:
        db.add(t)
    db.commit()
    return len(trades)


def insert_dividends(db: Session, dividends: list[Dividend]) -> int:
    for d in dividends:
        db.add(d)
    db.commit()
    return len(dividends)
